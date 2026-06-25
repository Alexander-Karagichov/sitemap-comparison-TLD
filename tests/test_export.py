import io

from openpyxl import load_workbook

from sitemap_compare.compare import Table
from sitemap_compare.export import to_xlsx


def _wb(data: bytes):
    return load_workbook(io.BytesIO(data))


def test_xlsx_has_comparison_and_duplicates_sheets():
    table = Table(
        headers=["brightdata.com", "bright.cn"],
        matched_rows=[["https://brightdata.com/a", "https://www.bright.cn/a"]],
        orphan_rows=[["", "https://www.bright.cn/x"]],
    )
    dup_rows = [("brightdata.com", "https://brightdata.com/a", 2)]
    wb = _wb(to_xlsx(table, dup_rows))
    assert wb.sheetnames == ["Comparison", "Duplicates"]


def test_xlsx_comparison_content():
    table = Table(
        headers=["brightdata.com", "bright.cn"],
        matched_rows=[["https://brightdata.com/a", "https://www.bright.cn/a"]],
        orphan_rows=[["", "https://www.bright.cn/x"]],
    )
    wb = _wb(to_xlsx(table, []))
    ws = wb["Comparison"]
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == ("brightdata.com", "bright.cn")           # header
    assert rows[1] == ("https://brightdata.com/a", "https://www.bright.cn/a")
    # orphan row: blank main cell (None or "") + tld url
    assert rows[2][0] in (None, "")
    assert rows[2][1] == "https://www.bright.cn/x"


def test_xlsx_duplicates_content():
    table = Table(headers=["a"], matched_rows=[], orphan_rows=[])
    dup_rows = [("brightdata.com", "https://brightdata.com/a", 3)]
    wb = _wb(to_xlsx(table, dup_rows))
    ws = wb["Duplicates"]
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == ("Site", "URL", "Count")
    assert rows[1] == ("brightdata.com", "https://brightdata.com/a", 3)
